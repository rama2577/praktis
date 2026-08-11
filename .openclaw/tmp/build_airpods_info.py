#!/usr/bin/env python3
"""Build 'AirPods 3 Info.xlsx' — price comparison across Amazon, eBay, Temu."""
import sys, os

XLSX_SKILL_DIR = os.path.expanduser("~/.openclaw-autoclaw/skills/xlsx")
for sub in [XLSX_SKILL_DIR, os.path.join(XLSX_SKILL_DIR, "templates")]:
    if sub not in sys.path:
        sys.path.insert(0, sub)

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from templates.base import (
    setup_sheet, style_header_row, style_data_row,
    font_body, font_caption, font_subheader, align_text, align_number,
    auto_fit_columns, CF_POSITIVE_FILL, CF_POSITIVE_FONT,
    NEUTRAL_600, PRIMARY, FORMATS,
)

OUT = "/Users/staff/Desktop/AirPods 3 Info.xlsx"
FX = 17823  # 1 USD ≈ IDR 17,823 (mid-market, 10 Aug 2026)

wb = Workbook()
wb.properties.creator = "Z.ai"

# ------------------------------------------------------------------
# Sheet 1: Price Comparison
# ------------------------------------------------------------------
ws = wb.active
ws.title = "Price Comparison"
LAST_COL = 7  # B..G
setup_sheet(ws, title="AirPods 3 (3rd Generation) \u2014 Price Comparison (10 Aug 2026)", last_col=LAST_COL)

headers = ["Source", "Product / Condition", "Price (IDR)", "Price (USD est.)", "Seller", "Notes"]
for col, h in enumerate(headers, start=2):
    ws.cell(row=4, column=col, value=h)
style_header_row(ws, 4, 2, LAST_COL)

rows = [
    # (source, product, idr, usd, seller, notes)
    ["eBay", "AirPods 3 (3rd Gen) \u2014 Pre-Owned", 1037847, 58, "smartsalesusa (99.5% positive)",
     "CHEAPEST genuine full set found (used condition)"],
    ["eBay", "AirPods 3 (3rd Gen) \u2014 Brand New", 1327966, 75, "raymond552 (0% feedback)",
     "Risky: brand-new seller with zero feedback history"],
    ["eBay", "AirPods 3 (3rd Gen) \u2014 Refurbished (Excellent)", 1503348, 84, "alphanuxenterprises", ""],
    ["eBay", "AirPods 3 (3rd Gen) \u2014 Open Box (MagSafe)", 2236946, 126, "\u2014", ""],
    ["Amazon", "AirPods 3 (3rd Gen) \u2014 new", "N/A", "N/A", "Amazon",
     "Discontinued \u2014 not available on Amazon US (only cases/accessories)"],
    ["Amazon", "AirPods 4 (current lineup)", 2190384, 123, "Amazon", "Reference: current-gen replacement model"],
    ["Amazon", "AirPods Pro 2 (current lineup)", 3347014, 188, "Amazon", "Reference model"],
    ["Amazon", "AirPods 2 (Renewed)", 1335422, 75, "Amazon", "Reference model (renewed)"],
    ["Temu", "AirPods 3 (3rd Gen) \u2014 genuine", "N/A", "N/A", "\u2014",
     "Browser blocked by Temu security verification; indexed Temu listings are non-genuine clones (~US$5\u20137) and accessories"],
]

r = 5
for i, row in enumerate(rows):
    for col, val in enumerate(row, start=2):
        ws.cell(row=r, column=col, value=val)
    style_data_row(ws, r, 2, LAST_COL, i)
    # alignment / formats
    ws.cell(row=r, column=3).alignment = align_number()
    ws.cell(row=r, column=3).number_format = FORMATS["integer"]
    ws.cell(row=r, column=4).alignment = align_number()
    if isinstance(row[2], (int, float)):
        ws.cell(row=r, column=4).number_format = FORMATS["currency_usd"]
    ws.cell(row=r, column=6).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.cell(row=r, column=5).alignment = align_text()
    ws.cell(row=r, column=2).alignment = align_text()
    r += 1

# Highlight the cheapest row (row 5)
for col in range(2, LAST_COL + 1):
    c = ws.cell(row=5, column=col)
    c.fill = CF_POSITIVE_FILL
    c.font = CF_POSITIVE_FONT

# Notes row under the table
note_r = r + 2
ws.cell(row=note_r, column=2, value="Notes:").font = font_subheader()
notes = [
    "Prices are as displayed by each site on 10 Aug 2026 (AutoClaw built-in browser).",
    "Amazon and eBay displayed prices in IDR (delivery location: Indonesia); Temu displays CAD on its Canada site.",
    "USD estimates converted at 1 USD = IDR 17,823 (mid-market rate, 10 Aug 2026).",
    "Temu could not be browsed: the site repeatedly served an image-based security verification (captcha) that blocks automated sessions.",
]
for i, n in enumerate(notes):
    c = ws.cell(row=note_r + 1 + i, column=2, value="\u2022 " + n)
    c.font = font_caption()
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)

ws.freeze_panes = "B5"
auto_fit_columns(ws, min_width=10, max_width=46, header_row=4, data_start_row=5)

# ------------------------------------------------------------------
# Sheet 2: Summary & Notes
# ------------------------------------------------------------------
ws2 = wb.create_sheet("Summary")
setup_sheet(ws2, title="Summary \u2014 Cheapest Option & Caveats", last_col=3)

rows2 = [
    ("Cheapest option", "eBay \u2014 AirPods 3 (3rd Gen) Pre-Owned at IDR 1,037,847 (\u2248US$58), seller smartsalesusa (99.5% positive)"),
    ("Amazon", "No genuine AirPods 3 available \u2014 model discontinued on Amazon US; current lineup is AirPods 4 / AirPods Pro 2"),
    ("Temu", "Not verifiable \u2014 security verification blocked the browser session; indexed Temu listings for 'AirPods 3' are non-genuine clones/accessories"),
    ("FX rate used", "1 USD \u2248 IDR 17,823 (mid-market, 10 Aug 2026 \u2014 XE / Bloomberg)"),
    ("Currency note", "Amazon & eBay showed IDR (delivery to Indonesia); Temu's Canada site shows CAD"),
]
r2 = 4
for i, (k, v) in enumerate(rows2):
    a = ws2.cell(row=r2, column=2, value=k)
    b = ws2.cell(row=r2, column=3, value=v)
    a.font = font_subheader()
    a.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    b.font = font_body()
    b.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws2.row_dimensions[r2].height = 34
    r2 += 1

r2 += 1
c = ws2.cell(row=r2, column=2, value="Caveats:")
c.font = font_subheader()
caveats = [
    "The eBay 'Brand New' listing (IDR 1,327,966) is from a seller with 0% feedback \u2014 verify seller reputation before buying.",
    "Pre-Owned price is the cheapest genuine full set found; condition is used, not new.",
    "Prices fluctuate; re-check before purchase. Temu comparison is based on web-search fallback because live browsing was blocked.",
]
for i, t in enumerate(caveats):
    cc = ws2.cell(row=r2 + 1 + i, column=2, value="\u2022 " + t)
    cc.font = font_caption()
    cc.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws2.row_dimensions[r2 + 1 + i].height = 28

r3 = r2 + len(caveats) + 2
ws2.cell(row=r3, column=2, value="Search URLs used:").font = font_subheader()
urls = [
    ("Amazon", "https://www.amazon.com/s?k=apple+airpods+3rd+generation"),
    ("eBay", "https://www.ebay.com/sch/i.html?_nkw=apple+airpods+3rd+generation"),
    ("Temu", "https://www.temu.com/search_result.html?search_key=airpods+3"),
]
for i, (name, url) in enumerate(urls):
    cc = ws2.cell(row=r3 + 1 + i, column=2, value=f"{name}:")
    cc.font = font_body()
    cc.alignment = align_text()
    u = ws2.cell(row=r3 + 1 + i, column=3, value=url)
    u.font = Font(name=font_body().name, size=9, color=PRIMARY)
    u.alignment = align_text()
    u.hyperlink = url

ws2.freeze_panes = "A5"
auto_fit_columns(ws2, min_width=10, max_width=70, header_row=4, data_start_row=5)

wb.save(OUT)
print("saved:", OUT)
