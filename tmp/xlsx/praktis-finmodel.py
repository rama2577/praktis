#!/usr/bin/env python3
"""Generate Analisa Bisnis Praktis — multi-sheet financial model (xlsx)."""
import sys, os

XLSX_SKILL_DIR = os.path.expanduser("~/.openclaw-autoclaw/skills/xlsx")
for sub in [XLSX_SKILL_DIR, os.path.join(XLSX_SKILL_DIR, "templates")]:
    if sub not in sys.path:
        sys.path.insert(0, sub)

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.comments import Comment
from templates.base import (
    FONT_NAME, HEADER_BOLD, PRIMARY, PRIMARY_LIGHT, NEUTRAL_900, NEUTRAL_600,
    NEUTRAL_200, NEUTRAL_100, NEUTRAL_50, NEUTRAL_0,
    ACCENT_POSITIVE, ACCENT_NEGATIVE, ACCENT_WARNING, CHART_COLORS,
)

OUT = os.path.expanduser("~/.openclaw-autoclaw/workspace/output/praktis-bisnis/Analisa Bisnis Praktis.xlsx")

# ── Finance number formats ──
FMT_RP   = 'Rp #,##0;(Rp #,##0);"-"'
FMT_RP0  = 'Rp #,##0;(Rp #,##0);"-"'
FMT_INT  = '#,##0;(#,##0);"-"'
FMT_PCT  = '0.0%;(0.0%);"-"'
FMT_2D   = '#,##0.00;(#,##0.00);"-"'

MONTHS = [f"Bulan {i}" for i in range(1, 13)]

# ── Style helpers ──
def style_header(cell, fill=PRIMARY, color="FFFFFF", bold=True):
    cell.font = Font(name=FONT_NAME, bold=bold, color=color, size=11)
    cell.fill = PatternFill("solid", fgColor=fill)
    cell.alignment = Alignment(horizontal="center", vertical="center")

def style_title(cell, size=14):
    cell.font = Font(name=FONT_NAME, bold=True, color=PRIMARY, size=size)

def style_section(cell, fill=PRIMARY_LIGHT, color=PRIMARY):
    cell.font = Font(name=FONT_NAME, bold=True, color=color, size=11)
    cell.fill = PatternFill("solid", fgColor=fill)

def style_input(cell):
    cell.font = Font(name=FONT_NAME, color="0000FF")  # blue = changeable
    cell.fill = PatternFill("solid", fgColor="FFF9C4")  # yellow bg = key assumption

def style_note(cell, text):
    cell.font = Font(name=FONT_NAME, size=8, italic=True, color=NEUTRAL_600)
    cell.value = text

wb = Workbook()
wb.properties.creator = "Z.ai"

# ═══════════════════════════════════════════════════════════
# SHEET 1 — ASUSMSI (Assumptions)
# ═══════════════════════════════════════════════════════════
ws = wb.active
ws.title = "Asumsi"
ws.sheet_view.showGridLines = False
ws.column_dimensions["A"].width = 4
ws.column_dimensions["B"].width = 46
ws.column_dimensions["C"].width = 18
ws.column_dimensions["D"].width = 18
ws.column_dimensions["E"].width = 18
ws.column_dimensions["F"].width = 40

ws["B2"] = "ASUMSI MODEL BISNIS — PRAKTIS (Direct-to-Market)"
style_title(ws["B2"], 15)
ws["B3"] = "Semua sel kuning bisa diubah — model menghitung ulang otomatis"
style_note(ws["B3"], "Sumber asumsi: dokumen analisa-bisnis-praktis-direct.md + riset pasar 14 Agu 2026")

r = 5
def section_row(text):
    global r
    ws.merge_cells(f"B{r}:E{r}")
    style_section(ws[f"B{r}"], PRIMARY, "FFFFFF")
    ws[f"B{r}"] = text
    r += 1

def input_row(label, value, fmt=FMT_RP, note=""):
    global r
    ws[f"B{r}"] = label
    ws[f"B{r}"].font = Font(name=FONT_NAME, color=NEUTRAL_900)
    c = ws[f"C{r}"]
    c.value = value
    c.number_format = fmt
    style_input(c)
    if note:
        ws[f"D{r}"] = note
        style_note(ws[f"D{r}"], note)
    r += 1

section_row("HARGA & PAKET")
input_row("Harga per transaksi (pay-per-use)", 250, FMT_RP, "Rekomendasi: Rp150-500 (di atas rekeningkoran.com Rp400/halaman)")
input_row("Paket Starter — kuota transaksi", 1000, FMT_INT)
input_row("Paket Starter — harga", 500000, FMT_RP)
input_row("Paket Growth — kuota transaksi", 3000, FMT_INT)
input_row("Paket Growth — harga", 1200000, FMT_RP)
input_row("Paket Enterprise — kuota transaksi", 10000, FMT_INT)
input_row("Paket Enterprise — harga", 3500000, FMT_RP)

section_row("BIAYA VARIABEL")
input_row("AI cost per transaksi (model routing)", 70, FMT_RP, "Asumsi GP 75-85% (MEMORY.md)")
input_row("Infra bulan 1-6 (Railway, dll)", 2000000, FMT_RP)
input_row("Infra bulan 7-12 (skala)", 4000000, FMT_RP)

section_row("ARPU & CHURN PER SKENARIO")
input_row("ARPU konservatif", 600000, FMT_RP)
input_row("ARPU moderat", 800000, FMT_RP)
input_row("ARPU optimis", 1000000, FMT_RP)
input_row("Churn konservatif (per bulan)", 0.06, FMT_PCT)
input_row("Churn moderat (per bulan)", 0.05, FMT_PCT)
input_row("Churn optimis (per bulan)", 0.04, FMT_PCT)

section_row("PELANGGAN BARU PER BULAN")
input_row("Konservatif — konstan", 3, FMT_INT)
input_row("Moderat — b1-4 / b5-8 / b9-12", 5, FMT_INT, "Dialokasikan bertahap di sheet skenario")
input_row("Optimis — b1-4 / b5-8 / b9-12", 10, FMT_INT)

section_row("MARKETING (Rp/bulan)")
input_row("Marketing fase 1 (b1-3) — konservatif", 2000000, FMT_RP)
input_row("Marketing fase 1 (b1-3) — moderat", 2000000, FMT_RP)
input_row("Marketing fase 1 (b1-3) — optimis", 3000000, FMT_RP)
input_row("Marketing fase 2 (b4-6) — konservatif", 5000000, FMT_RP)
input_row("Marketing fase 2 (b4-6) — moderat", 10000000, FMT_RP)
input_row("Marketing fase 2 (b4-6) — optimis", 15000000, FMT_RP)
input_row("Marketing fase 3 (b7-12) — konservatif", 8000000, FMT_RP)
input_row("Marketing fase 3 (b7-12) — moderat", 16000000, FMT_RP)
input_row("Marketing fase 3 (b7-12) — optimis", 25000000, FMT_RP)

section_row("INVESTASI AWAL (one-off)")
input_row("Fitur direct-market (dev 6-9 minggu)", 25000000, FMT_RP, "Payment, self-serve, export CSV, landing page")
input_row("Domain, branding, tooling", 5000000, FMT_RP)
input_row("Dana cadangan (buffer)", 10000000, FMT_RP)

ws.freeze_panes = "A5"
ws.sheet_view.zoomScale = 100

# ═══════════════════════════════════════════════════════════
# SCENARIO SHEET BUILDER
# ═══════════════════════════════════════════════════════════
def build_scenario(name, arpu_ref, churn_ref, new_customers, marketing_schedule, ai_pct=0.14):
    """
    new_customers: list 12 angka pelanggan baru per bulan
    marketing_schedule: list 12 angka marketing per bulan
    """
    ws2 = wb.create_sheet(name)
    ws2.sheet_view.showGridLines = False
    ws2.column_dimensions["A"].width = 14
    for i in range(2, 15):
        ws2.column_dimensions[get_column_letter(i)].width = 15

    # Title
    ws2.merge_cells("A1:N1")
    ws2["A1"] = f"PROYEKSI 12 BULAN — {name.upper()}"
    style_title(ws2["A1"], 14)

    headers = ["Bulan", "Pelanggan Awal", "Pelanggan Baru", "Churn", "Pelanggan Akhir",
               "MRR", "Revenue", "AI Cost", "Infra", "Marketing", "Total Opex",
               "Net Cash Flow", "Kumulatif", "CAC Bulanan"]
    for i, h in enumerate(headers):
        cell = ws2.cell(row=3, column=i + 1, value=h)
        style_header(cell)

    # Rows 4..15 = bulan 1..12
    for m in range(12):
        row = 4 + m
        ws2.cell(row=row, column=1, value=m + 1).number_format = FMT_INT
        # Pelanggan Awal
        if m == 0:
            ws2.cell(row=row, column=2, value=0)
        else:
            ws2.cell(row=row, column=2, value=f"=L{row-1}")
        # Pelanggan Baru
        ws2.cell(row=row, column=3, value=new_customers[m])
        # Churn
        ws2.cell(row=row, column=4, value=f"=IF(B{row}=0,0,ROUND(B{row}*{churn_ref},0))")
        # Pelanggan Akhir
        ws2.cell(row=row, column=5, value=f"=B{row}+C{row}-D{row}")
        # MRR
        ws2.cell(row=row, column=6, value=f"=E{row}*{arpu_ref}")
        # Revenue = MRR
        ws2.cell(row=row, column=7, value=f"=F{row}")
        # AI Cost
        ws2.cell(row=row, column=8, value=f"=ROUND(G{row}*{ai_pct},0)")
        # Infra
        infra_val = 2000000 if m < 6 else 4000000
        ws2.cell(row=row, column=9, value=infra_val)
        # Marketing
        ws2.cell(row=row, column=10, value=marketing_schedule[m])
        # Total Opex
        ws2.cell(row=row, column=11, value=f"=H{row}+I{row}+J{row}")
        # Net Cash Flow
        ws2.cell(row=row, column=12, value=f"=G{row}-K{row}")
        # Kumulatif
        if m == 0:
            ws2.cell(row=row, column=13, value=f"=L{row}")
        else:
            ws2.cell(row=row, column=13, value=f"=M{row-1}+L{row}")
        # CAC = marketing / pelanggan baru
        ws2.cell(row=row, column=14, value=f'=IF(C{row}=0,"-",J{row}/C{row})')

        for col in range(1, 15):
            cell = ws2.cell(row=row, column=col)
            if col in (6, 7, 8, 9, 10, 11, 12, 13, 14):
                cell.number_format = FMT_RP if col != 14 else FMT_RP0
            elif col in (2, 3, 4, 5):
                cell.number_format = FMT_INT
            if row % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=NEUTRAL_100)
            cell.font = Font(name=FONT_NAME, color=NEUTRAL_900)

    # Break-even row
    br = 17
    ws2.cell(row=br, column=1, value="Break-even (kumulatif >= 0)").font = Font(name=FONT_NAME, bold=True, color=PRIMARY)
    ws2.cell(row=br, column=2, value='=IFERROR(MATCH(TRUE,INDEX(M4:M15>=0,0),0),"-")')
    ws2.cell(row=br, column=2).number_format = FMT_INT
    ws2.cell(row=br, column=2).font = Font(name=FONT_NAME, bold=True, color=ACCENT_POSITIVE)

    br2 = br + 1
    ws2.cell(row=br2, column=1, value="Total pelanggan akhir (bulan 12)").font = Font(name=FONT_NAME, color=NEUTRAL_600)
    ws2.cell(row=br2, column=2, value="=E15")
    ws2.cell(row=br2, column=2).number_format = FMT_INT

    br3 = br + 2
    ws2.cell(row=br3, column=1, value="Revenue kumulatif 12 bulan").font = Font(name=FONT_NAME, color=NEUTRAL_600)
    ws2.cell(row=br3, column=2, value="=SUM(G4:G15)")
    ws2.cell(row=br3, column=2).number_format = FMT_RP

    ws2.freeze_panes = "A4"
    return ws2

# ── Marketing schedules ──
def mkt(phase1, phase2, phase3):
    return [phase1]*3 + [phase2]*3 + [phase3]*6

mkt_kons = mkt(2000000, 5000000, 8000000)
mkt_mod  = mkt(2000000, 10000000, 16000000)
mkt_opt  = mkt(3000000, 15000000, 25000000)

# ── New customers schedules ──
new_kons = [3]*12
new_mod  = [5,5,5,5, 8,8,8,8, 12,12,12,12]
new_opt  = [10,10,10,10, 18,18,18,18, 25,25,25,25]

ws_kons = build_scenario("Konservatif", "Asumsi!$C$15", "Asumsi!$C$18", new_kons, mkt_kons)
ws_mod  = build_scenario("Moderat", "Asumsi!$C$16", "Asumsi!$C$19", new_mod, mkt_mod)
ws_opt  = build_scenario("Optimis", "Asumsi!$C$17", "Asumsi!$C$20", new_opt, mkt_opt)

# ═══════════════════════════════════════════════════════════
# SHEET — UNIT ECONOMICS
# ═══════════════════════════════════════════════════════════
ws3 = wb.create_sheet("Unit Economics")
ws3.sheet_view.showGridLines = False
for col, w in zip("ABCDEF", [40, 20, 20, 20, 20, 30]):
    ws3.column_dimensions[col].width = w

ws3["A1"] = "UNIT ECONOMICS (SKENARIO MODERAT)"
style_title(ws3["A1"], 14)

rows = [
    ("ARPU (Rp/bulan)", "=Asumsi!$C$16", FMT_RP),
    ("Churn bulanan", "=Asumsi!$C$19", FMT_PCT),
    ("Gross margin (1 - AI cost %)", "=1-0.14", FMT_PCT),
    ("CAC (Rp, rata-rata)", "=AVERAGE(Moderat!N4:N15)", FMT_RP),
    ("LTV (Rp)", "=IFERROR(B2*B4/B3,0)", FMT_RP),
    ("LTV : CAC", '=IFERROR(B5/B6,0)&"x"', None),
    ("Payback (bulan)", "=IFERROR(B6/(B2*B4),0)", FMT_2D),
    ("MRR bulan 12", "=Moderat!F15", FMT_RP),
    ("Revenue kumulatif 12 bln", "=Moderat!G16", FMT_RP),
]
ws3["A3"] = "Metrik"; ws3["B3"] = "Nilai"; ws3["C3"] = "Benchmark"
for c in ("A3", "B3", "C3"):
    style_header(ws3[c])
for i, (label, formula, fmt) in enumerate(rows):
    r = 4 + i
    ws3[f"A{r}"] = label
    ws3[f"A{r}"].font = Font(name=FONT_NAME, color=NEUTRAL_900)
    ws3[f"B{r}"] = formula
    if fmt:
        ws3[f"B{r}"].number_format = fmt
    ws3[f"B{r}"].font = Font(name=FONT_NAME, bold=True, color=PRIMARY)

benchmarks = ["", "Rp 600-800rb", "3-6%", "70-85%", "Rp 400-800rb", "Rp 10-15 jt", "> 3x", "1-3 bln", "", ""]
for i, b in enumerate(benchmarks):
    if b:
        ws3[f"C{4+i}"] = b
        style_note(ws3[f"C{4+i}"], b)

ws3["A15"] = "Catatan: LTV:CAC di atas 3x dianggap sehat. Payback < 3 bulan sangat baik untuk self-serve SaaS."
style_note(ws3["A15"], "")

# ═══════════════════════════════════════════════════════════
# SHEET — RINGKASAN (Dashboard)
# ═══════════════════════════════════════════════════════════
ws4 = wb.create_sheet("Ringkasan")
ws4.sheet_view.showGridLines = False
ws4.column_dimensions["A"].width = 38
for i in range(2, 7):
    ws4.column_dimensions[get_column_letter(i)].width = 20

ws4["A1"] = "RINGKASAN ANALISA BISNIS — PRAKTIS"
style_title(ws4["A1"], 16)
ws4["A2"] = "AI Bookkeeping: PDF/foto/Excel/CSV → draft jurnal + laporan keuangan · Agustus 2026"
style_note(ws4["A2"], "")

ws4["A4"] = "Metrik Kunci"; ws4["B4"] = "Konservatif"; ws4["C4"] = "Moderat"; ws4["D4"] = "Optimis"
for c in ("A4", "B4", "C4", "D4"):
    style_header(ws4[c])

metrics = [
    ("Investasi awal (Rp)", "=SUM(Asumsi!C28:C30)", "=SUM(Asumsi!C28:C30)", "=SUM(Asumsi!C28:C30)", FMT_RP),
    ("Pelanggan akhir (bln 12)", "=Konservatif!E15", "=Moderat!E15", "=Optimis!E15", FMT_INT),
    ("MRR bulan 12 (Rp)", "=Konservatif!F15", "=Moderat!F15", "=Optimis!F15", FMT_RP),
    ("Revenue kumulatif 12 bln", "=Konservatif!G16", "=Moderat!G16", "=Optimis!G16", FMT_RP),
    ("Net cash kumulatif (bln 12)", "=Konservatif!M15", "=Moderat!M15", "=Optimis!M15", FMT_RP),
    ("Break-even (bulan ke-)", "=Konservatif!B17", "=Moderat!B17", "=Optimis!B17", FMT_INT),
    ("CAC rata-rata (Rp)", "=AVERAGE(Konservatif!N4:N15)", "=AVERAGE(Moderat!N4:N15)", "=AVERAGE(Optimis!N4:N15)", FMT_RP),
]
for i, (label, f1, f2, f3, fmt) in enumerate(metrics):
    r = 5 + i
    ws4[f"A{r}"] = label
    ws4[f"A{r}"].font = Font(name=FONT_NAME, color=NEUTRAL_900)
    for j, f in enumerate((f1, f2, f3)):
        cell = ws4.cell(row=r, column=2 + j, value=f)
        cell.number_format = fmt
        cell.font = Font(name=FONT_NAME, bold=True, color=PRIMARY)
    if r % 2 == 1:
        for j in range(4):
            ws4.cell(row=r, column=1 + j).fill = PatternFill("solid", fgColor=NEUTRAL_100)

# Chart 1: Revenue per bulan (bar)
chart = BarChart()
chart.type = "col"
chart.title = "Revenue per Bulan (3 Skenario)"
chart.y_axis.title = "Rp"
chart.x_axis.title = "Bulan"
chart.height = 9
chart.width = 22
for wsx, col in ((ws_kons, "B"), (ws_mod, "C"), (ws_opt, "D")):
    # data: revenue rows 4..15 (G col) → but chart needs per-sheet refs
    pass
# Use explicit refs per sheet
data_k = Reference(ws_kons, min_col=7, min_row=3, max_row=15)
data_m = Reference(ws_mod, min_col=7, min_row=3, max_row=15)
data_o = Reference(ws_opt, min_col=7, min_row=3, max_row=15)
cats = Reference(ws_kons, min_col=1, min_row=4, max_row=15)
for data, title in ((data_k, "Konservatif"), (data_m, "Moderat"), (data_o, "Optimis")):
    chart.add_data(data, titles_from_data=True)
    chart.series[-1].tx = None  # will fix titles below
chart.set_categories(cats)
# rename series via explicit series titles
chart.series[0].tx = Reference(ws4, min_col=2, min_row=4)
chart.series[1].tx = Reference(ws4, min_col=3, min_row=4)
chart.series[2].tx = Reference(ws4, min_col=4, min_row=4)
ws4.add_chart(chart, "A16")

# Chart 2: Net cash kumulatif (line)
line = LineChart()
line.title = "Net Cash Flow Kumulatif (3 Skenario)"
line.y_axis.title = "Rp"
line.height = 9
line.width = 22
lk = Reference(ws_kons, min_col=13, min_row=3, max_row=15)
lm = Reference(ws_mod, min_col=13, min_row=3, max_row=15)
lo = Reference(ws_opt, min_col=13, min_row=3, max_row=15)
for data in (lk, lm, lo):
    line.add_data(data, titles_from_data=True)
line.set_categories(cats)
line.series[0].tx = Reference(ws4, min_col=2, min_row=4)
line.series[1].tx = Reference(ws4, min_col=3, min_row=4)
line.series[2].tx = Reference(ws4, min_col=4, min_row=4)
ws4.add_chart(line, "A33")

# ═══════════════════════════════════════════════════════════
# SHEET — MARKETING PLAN
# ═══════════════════════════════════════════════════════════
ws5 = wb.create_sheet("Marketing Plan")
ws5.sheet_view.showGridLines = False
for col, w in zip("ABCDEFG", [10, 18, 44, 22, 18, 18, 26]):
    ws5.column_dimensions[col].width = w

ws5["A1"] = "RENCANA PEMASARAN 12 BULAN (SKENARIO MODERAT)"
style_title(ws5["A1"], 14)

headers = ["Fase", "Bulan", "Aktivitas", "Budget", "Target Pelanggan", "CAC Harapan", "Catatan"]
for i, h in enumerate(headers):
    style_header(ws5.cell(row=3, column=i + 1, value=h))

plan = [
    ("Fase 1", "1-3", "Konten edukasi (LinkedIn, blog), SEO tutorial rekening koran, komunitas akuntan/KAP, cold outreach 200 target", 6000000, "10-15", "Rp 400-500rb", "Validasi product-market fit, CAC rendah"),
    ("Fase 2", "4-6", "Google Ads (keyword: rekening koran ke excel/jurnal), Meta Ads retargeting, referral program, webinar bulanan", 30000000, "20", "Rp 600-700rb", "Scale kanal berbayar"),
    ("Fase 3", "7-12", "Google + Meta + YouTube, partnership KAP/komunitas/kampus, 1 sales AE", 96000000, "50+", "Rp 700-800rb", "Sales-assisted untuk enterprise"),
]
r = 4
for row in plan:
    for j, val in enumerate(row):
        ws5.cell(row=r, column=j + 1, value=val)
    ws5.cell(row=r, column=4).number_format = FMT_RP
    r += 1

ws5[f"B{r}"] = "TOTAL"
ws5[f"B{r}"].font = Font(name=FONT_NAME, bold=True, color=PRIMARY)
ws5[f"D{r}"] = "=SUM(D4:D6)"
ws5[f"D{r}"].number_format = FMT_RP
ws5[f"D{r}"].font = Font(name=FONT_NAME, bold=True, color=PRIMARY)
ws5[f"E{r}"] = "80-85"
ws5[f"G{r}"] = "≈ Rp 132 jt sesuai dokumen analisa"
style_note(ws5[f"G{r}"], "≈ Rp 132 jt")

# ═══════════════════════════════════════════════════════════
# QA + save
# ═══════════════════════════════════════════════════════════
os.makedirs(os.path.dirname(OUT), exist_ok=True)
wb.save(OUT)
print(f"✅ Saved: {OUT}")
